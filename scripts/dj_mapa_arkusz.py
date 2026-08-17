"""Arkusz mapy DJ-ów — Audioriver + Garbicz, line-up 2026 plus archiwa.

Puste pole znaczy „nie znalazłam", nie „nie ma". Nigdzie nie ma wartości
zgadniętej: link wpisany jest linkiem, który widziałam, a nie takim, który
prawdopodobnie istnieje (ADR-005). Kolumna `uwagi` mówi, czego brakuje i czemu.

KOLUMNY Z IMIENIEM I NAZWISKIEM NIE MA — decyzja Janka po pilocie. Na dwudziestu
artystach żaden nie podawał nazwiska w swoim bio; jedyne, co się znajdowało, to
cudze posty i artykuły, a sklejanie z nich prawdziwych nazwisk jest zbieraniem
danych osobowych, nie research'em zawodowym. Kolumna została usunięta zamiast
zostawiona pustą, żeby nikt jej później nie „uzupełnił".

Trzy źródła, każde z własnym stopniem pewności:

  * FESTIWALE — z list line-upów, pewne.
  * APPLE MUSIC + UTWORY — z API iTunes, dopasowanie po dokładnej nazwie.
    Nazwa nie wystarcza przy krótkich ksywach, więc wiersz z gatunkiem spoza
    muzyki klubowej („Nilo — K-Pop", „Mimi — Anime") dostaje ostrzeżenie
    zamiast cichego linku do kogoś zupełnie innego.
  * MIKSY — z trzech miejsc, każde o innym charakterze: 1001Tracklists
    (baza z datą i miejscem), archiwa festiwalowe ze SoundCloud (kuratorowane
    playlisty rocznikowe Garbicza + wyszukiwarka api-v2 dla Audioriver, który
    takich playlist nie ma) oraz YouTube, gdzie każdy wynik wymaga oceny.
  * SOUNDCLOUD — w większości ubocznie z archiwów: konto wrzucające set jest
    kontem artysty w połowie przypadków, i to wystarcza za pewne dopasowanie.

KOLUMNY INSTAGRAM NIE MA — decyzja Janka po ~80 artystach: do słuchania cudzych
setów Instagram jest bezużyteczny, liczy się SoundCloud i YouTube. Zebrane
uchwyty zostają w `socials.json`, gdyby kiedyś wróciły; z arkusza wypadają.

KOLUMN REZYDENCJA I AFILIACJA TEŻ NIE MA — decyzja Janka 2026-08-14. Przy ponad
tysiącu artystów wypełnione były 24 wiersze, bo to jedyne pole bez źródła, które da
się zaciągnąć hurtem: każdy wpis wymagał osobnego wejścia na stronę klubu.
Kolumna, która w 98% mówi „nie wiem", uczy ignorować całą tabelę. Wartości
leżą w `socials.json` i wrócą, jeśli znajdzie się sposób masowy.
"""

from __future__ import annotations

import json
import pathlib

OUT = pathlib.Path("experiments_priv/2026-08-03_dj_mapa")

# ── PALETA ──────────────────────────────────────────────────────────────────
# Kolor niesie ZNACZENIE, nie ozdobę: jedna rodzina barw = jeden wymiar.
# Miejsca z publicznością idą w ciepłe i nasycone, nagrania bez publiczności
# w chłodne i blade — bo to jest podział, który najbardziej zmienia sposób
# grania. Rola w programie dostaje osobną skalę dobową: świt jasny, noc ciemna.
KOLORY = {
    "typ": {
        "festiwal":  "C6E0B4",   # zieleń — plener, tłum
        "klub":      "D9C2E9",   # fiolet — noc, wnętrze
        "warehouse": "D0CECE",   # szarość — surowa przestrzeń
        "rave":      "F4B8D8",   # magenta — miejsce nieoczywiste
        "plener":    "E2EFDA",   # jasna zieleń
        "radio":     "BDD7EE",   # błękit — bez publiczności
        "studio":    "DEEBF7",   # jaśniejszy błękit
        "podcast":   "D6E4F0",   # najbledszy błękit
        "stream":    "EDF3F8",
    },
    "format": {
        "dj-set": "FFFFFF",      # domyślny — bez wyróżnienia
        "live":   "FFD966",      # bursztyn — gra z maszyn
        "b2b":    "FFE699",      # jaśniejszy bursztyn — dwoje ludzi
        "winyl":  "E6D3B3",      # beż — płyty
    },
    "rola": {                    # skala dobowa: od świtu do nocy
        "wschod-slonca": "FFF2CC",
        "poranek":       "FCE4D6",
        "otwarcie":      "E2EFDA",
        "popoludnie":    "FFE0B2",   # dołożone przy rocznikach Garbicza
        "zachod-slonca": "FFC08A",   # — 34 sety opisane wprost jako te pory
        "peak":          "FF9999",
        "noc":           "B4C7E7",
        "afterhour":     "9DC3E6",
        "zamkniecie":    "C9A0DC",
        "all-night":     "8EA9DB",
    },
    "zrodlo": {
        "1001tracklists": "C6E0B4",   # baza z datą i miejscem — najpewniejsze
        "soundcloud":     "FFE699",   # wrzut artysty — pewne, ale bez metadanych
        "youtube":        "F8CBAD",   # wymaga oceny (składanki fanowskie)
    },
    "pewnosc": {
        "potwierdzone": "C6E0B4",
        "niepewne":     "FFD966",
    },
}


# Zebrane ręcznie w pilocie: instagram, soundcloud, rezydencja/afiliacja, uwagi.
RECZNIE = {
 "Ben Klock": ("ben_klock", "ben-klock", "Berghain (Berlin)", ""),
 "Len Faki": ("len_faki", "lenfaki", "Berghain (Berlin)", ""),
 "Ellen Allien b2b Salome": ("ellen.allien", "ellen-allien",
    "BPitch / UFO Inc. (wytwórnie)", "rezydencji klubowej nie podaje"),
 "Mano Le Tough": ("manoletough", "manoletough", "Maeve (wytwórnia)",
    "rezydencji klubowej nie podaje"),
 "I Hate Models": ("ihatemodels1", "IHATEMODELS", "Disco Inferno (wytwórnia)",
    "rezydencji klubowej nie podaje"),
 "Rødhåd": ("rodhad_wsnwg", "rodhad", "WSNWG / Dystopian (wytwórnie)",
    "drugi profil @rodhad_dystopian — aktywny jest @rodhad_wsnwg"),
 "Palms Trax": ("palmstrax", "palmstrax", "", "rezydencji nie znalazłam"),
 "Cassy": ("cassyofficial", "c-a-s-s-y-48501209", "Kwench Records (wytwórnia)", ""),
 "Kevin Saunderson": ("kevinsaunderson", "kevinsaunderson", "Inner City / KMS", ""),
 "Roman Flügel": ("roman_fluegel", "", "", "SoundCloud nie znaleziony"),
 "Paramida": ("paramidaaa", "paramida",
    "Panorama Bar (Berlin) · Love On The Rocks", ""),
 "Catz 'n Dogz": ("catz_n_dogz", "catzndogz", "Pets Recordings (Szczecin)",
    "duet: Greg & Voitek"),
 "Héctor Oaks": ("hectoroaks", "hectoroaks",
    "Herrensauna (Berlin) · Bassiani (Tbilisi)", ""),
 "Gene On Earth": ("gene_on_earth", "geneonearth", "Limousine Dream (wytwórnia)", ""),
 "Ash Lauryn": ("ash_lauryn_", "ashlauryn", "NTS Radio (rezydencja radiowa)",
    "platforma Underground & Black"),
 "Cinthie": ("cinthie_dj", "cinthie",
    "Robert Johnson (Offenbach) · Public Records (NYC)", "803 Crystal Grooves"),
 "Fort Romeau": ("fortromeau", "fortromeau", "", "rezydencji nie znalazłam"),
 "D. Tiffany": ("", "", "Planet Euphorique (wytwórnia)",
    "Instagram NIEPEWNY — wyszukiwarka zwróciła tylko wytwórnię"),
 "Praktyczna Pani": ("djpraktycznapani", "", "Full House Group (head of music)", ""),
 "Błażej Malinowski": ("blazejmalinowski", "blazejmalinowskiofficial",
    "Inner Tension (wytwórnia)", ""),
}


def main() -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fest = json.loads((OUT / "festiwale.json").read_text())
    soc = {}
    if (OUT / "socials.json").exists():
        soc = json.loads((OUT / "socials.json").read_text())
    # Apple leci partiami, w kolejności, w jakiej rosła baza: line-up 2026,
    # potem roczniki Garbicza, potem Audioriver, potem reszta. Osobne pliki,
    # bo jeden przebieg nadpisałby drugi.
    apple = json.loads((OUT / "apple.json").read_text())
    for dodatek in ("apple_roczniki.json", "apple_audioriver.json",
                    "apple_reszta.json", "apple_wisloujscie.json"):
        if (OUT / dodatek).exists():
            extra = json.loads((OUT / dodatek).read_text())
            apple["artysci"].extend(extra["artysci"])
            apple["utwory"].extend(extra["utwory"])
    ap = {r["ksywa"]: r for r in apple["artysci"]}
    # Bandcamp — decyzja Janka 2026-08-14: „jak ktoś nie wrzuca na Apple Music
    # to sprawdźmy Bandcampa". Brak w Apple nie znaczy brak muzyki: przy
    # line-upie Wisłoujścia, wyłącznie polskim i undergroundowym, Apple trafiło
    # 14%, bo ci ludzie wydają na Bandcampie i winylu.
    bc, bc_plyty = {}, []
    if (OUT / "bandcamp.json").exists():
        b = json.loads((OUT / "bandcamp.json").read_text())
        bc = {r["ksywa"]: r for r in b["artysci"]}
        bc_plyty = b["wydawnictwa"]
    # Resident Advisor jako globalny szkielet (decyzja Janka 2026-08-14).
    # Wraca tu kolumna WYTWÓRNIE, usunięta tego samego dnia — bo wróciło jej
    # źródło: RA podaje je strukturalnie, dziesięć dla Bena Klocka w jednym
    # zapytaniu, zamiast jednego wpisu na pięćdziesiąt wejść na stronę klubu.
    rap, wystepy = {}, []
    if (OUT / "ra.json").exists():
        r = json.loads((OUT / "ra.json").read_text())
        rap = {x["ksywa"]: x for x in r["artysci"]}
        wystepy = r["wystepy"]
    miksy = []
    if (OUT / "miksy.json").exists():
        miksy = json.loads((OUT / "miksy.json").read_text())

    wb = Workbook()
    naglowek = Font(bold=True, color="FFFFFF")
    tlo = PatternFill("solid", fgColor="333333")
    ostrzezenie = PatternFill("solid", fgColor="FFF2CC")

    # Tytuły z SoundCloud niosą znaki sterujące, których Excel nie przyjmuje
    # („CØSMIC VØYAGE episode 15. 물고기들의 꿈"). Czyścimy w jednym miejscu,
    # zamiast łatać każdą kolumnę osobno — i tylko znaki sterujące, bo koreański
    # i skandynawskie „Ø" są poprawną częścią nazwy.
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    def czysto(v):
        return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

    def arkusz(ws, kolumny, wiersze, szer):
        wiersze = [[czysto(c) for c in w] for w in wiersze]
        ws.append(kolumny)
        for c in ws[1]:
            c.font, c.fill = naglowek, tlo
            c.alignment = Alignment(vertical="center")
        for w in wiersze:
            ws.append(w)
        # get_column_letter, a nie stały alfabet — arkusz Miksy przekroczył
        # kolumnę J i szerokości po cichu przestawały się stosować.
        for i, s in enumerate(szer, start=1):
            ws.column_dimensions[get_column_letter(i)].width = s
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    ws = wb.active
    ws.title = "Artyści"
    # Line-up 2026 plus artyści, którzy wyszli z archiwów obu festiwali
    # (Garbicz 2013-2026, Audioriver 2009-2026). Ci drudzy mają puste pole
    # „festiwal 2026" — grali kiedyś, niekoniecznie w tym roku, i to
    # rozróżnienie musi być widoczne.
    miksow: dict[str, list] = {}
    lata: dict[str, set] = {}
    for m in miksy:
        k = m.get("ksywa")
        if not k:
            continue
        miksow.setdefault(k, []).append(m)
        if "garbicz" in (m.get("wydarzenie") or "").lower() and m.get("data"):
            lata.setdefault(k, set()).add(str(m["data"])[:4])

    znani = {v["ksywa"] for v in fest.values()}
    z_rocznikow = sorted({m["ksywa"] for m in miksy
                          if m.get("ksywa") and m["ksywa"] not in znani},
                         key=str.lower)
    pozycje = list(fest.values()) + [
        {"ksywa": k, "wystapienia": []} for k in z_rocznikow]

    wiersze, podejrzane_wiersze = [], []
    for i, v in enumerate(pozycje, start=2):
        ksywa = v["ksywa"]
        a = ap.get(ksywa, {})
        _, sc0, _, uw0 = RECZNIE.get(ksywa, ("", "", "", ""))
        e = soc.get(ksywa, {})
        sc = e.get("soundcloud") or sc0
        uw = e.get("uwagi") or uw0
        kand = e.get("kandydaci") or []
        b_ = bc.get(ksywa) or {}
        if b_.get("to_wytwornia"):
            uw = (uw + " · " if uw else "") + (
                "Bandcamp oznacza ten profil jako WYTWÓRNIĘ, nie artystę")
        if b_.get("do_sprawdzenia"):
            uw = (uw + " · " if uw else "") + (
                f"Bandcamp: gatunek „{b_.get('gatunek_bandcamp') or 'brak'}" + "” "
                "— sprawdź, czy to ten sam człowiek")
            podejrzane_wiersze.append(i)
        if a.get("do_sprawdzenia"):
            uw = (uw + " · " if uw else "") + (
                f"Apple: gatunek „{a.get('gatunek_apple') or 'brak'}" + "” "
                "— prawdopodobnie INNY wykonawca o tej samej nazwie")
            podejrzane_wiersze.append(i)
        wiersze.append([
            ksywa,
            (sc if str(sc).startswith("http") else f"https://soundcloud.com/{sc}") if sc else "",
            a.get("apple_music") or "",
            a.get("gatunek_apple") or "",
            (bc.get(ksywa) or {}).get("bandcamp") or "",
            (bc.get(ksywa) or {}).get("lokalizacja") or "",
            (bc.get(ksywa) or {}).get("gatunek_bandcamp") or "",
            (rap.get(ksywa) or {}).get("wytwornie") or "",
            (rap.get(ksywa) or {}).get("kraj") or "",
            (rap.get(ksywa) or {}).get("obserwujacych") or "",
            (rap.get(ksywa) or {}).get("wystepow_przeszlych") or "",
            " · ".join(v["wystapienia"]),
            len(miksow.get(ksywa, [])),
            " ".join(sorted(lata.get(ksywa, set()))),
            uw,
            "\n".join(kand),
        ])
    # Rezydencja i afiliacja wypadły 2026-08-14 (decyzja Janka). Były jedynym
    # polem bez taniego źródła — 24 wypełnione na 1007 — więc kolumna głównie
    # ogłaszała, czego nie wiemy. Zebrane wartości zostają w socials.json.
    arkusz(ws, ["ksywa sceniczna", "SoundCloud", "Apple Music", "gatunek wg Apple",
                "Bandcamp", "skąd (wg Bandcamp)", "gatunek wg Bandcamp",
                "wytwórnie (RA)", "kraj (RA)", "obserwujących (RA)",
                "występów w RA", "festiwal 2026", "miksów w bazie",
                "lata w Garbiczu", "uwagi", "kandydaci do rozstrzygnięcia"],
           wiersze, [26, 40, 50, 16, 42, 26, 18, 46, 16, 15, 14, 24, 14, 26, 54, 64])
    for r in podejrzane_wiersze:
        for c in range(1, 17):
            ws.cell(row=r, column=c).fill = ostrzezenie

    ws2 = wb.create_sheet("Utwory")
    arkusz(ws2, ["ksywa", "tytuł", "album", "rok", "link", "źródło"],
           [[t["ksywa"], t["tytul"], t["album"], t["rok"], t["link"], t["zrodlo"]]
            for t in apple["utwory"]]
           + [[t["ksywa"], t["tytul"], t["typ"], "", t["link"], t["zrodlo"]]
              for t in bc_plyty],
           [26, 44, 44, 8, 60, 40])

    ws3 = wb.create_sheet("Miksy")
    # `opis` idzie na koniec, bo bywa akapitem — to jest najcenniejsza kolumna
    # (kto grał przed kim, o której, w jakiej pogodzie), ale zjadłaby widok
    # gdyby stała między polami słownikowymi.
    arkusz(ws3, ["ksywa", "tytuł", "wydarzenie", "typ", "scena", "format",
                 "rola", "czas", "data", "źródło", "pewność", "długość (min)",
                 "konto (wrzucił)", "kto grał obok", "link", "opis od artysty"],
           [[m["ksywa"], m["tytul"], m.get("wydarzenie", ""), m.get("typ", ""),
             m.get("scena", ""), m.get("format", ""), m.get("rola", ""),
             m.get("czas", ""), m.get("data", ""), m.get("zrodlo", ""),
             m.get("pewnosc", ""), m.get("dlugosc_min", ""), m.get("konto", ""),
             m.get("sasiedztwo", ""), m["link"], m.get("opis", "")]
            for m in miksy],
           [24, 50, 18, 12, 16, 10, 14, 20, 12, 14, 12, 12, 22, 40, 56, 90])
    for r, mm in enumerate(miksy, start=2):
        for pole, kolumna in (("typ", 4), ("format", 6), ("rola", 7),
                              ("zrodlo", 10), ("pewnosc", 11)):
            barwa = KOLORY.get(pole, {}).get(mm.get(pole, ""))
            if barwa and barwa != "FFFFFF":
                ws3.cell(row=r, column=kolumna).fill = PatternFill("solid", fgColor=barwa)
    if not miksy:
        ws3["A2"] = "jeszcze nie zebrane — miksy z YouTube zbierane partiami"

    # PROGRAMY siedzą osobno od MIKSÓW, bo to dwie różne rzeczy:
    #   miks  = nagranie, które da się odtworzyć (kolumna `link` wypełniona
    #           w 100% i to jest cały sens tamtej tabeli);
    #   slot  = punkt w rozkładzie — kto, gdzie, od której do której.
    # Wrzucenie slotów do Miksów zepsułoby znaczenie kolumny `link`.
    #
    # Zakładka jest OGÓLNA, nie „Wisłoujście": pierwsza kolumna mówi, czyj to
    # program. Dziś mamy tylko jeden, bo tylko Wisłoujście publikuje pełny
    # timetable w pliku — Garbicz i Audioriver dołożą się tu bez przebudowy,
    # gdy ich rozkłady da się zdobyć.
    programy = []
    for plik in ("wisloujscie_program.json",):
        if (OUT / plik).exists():
            programy += json.loads((OUT / plik).read_text())
    if programy:
        wsp = wb.create_sheet("Programy")
        arkusz(wsp, ["festiwal", "ksywa", "dzień", "data", "scena",
                     "charakter sceny", "start", "koniec", "rola", "format"],
               [[p.get("wydarzenie", ""), p["ksywa"], p["dzien"], p["data"],
                 p["scena"], p["charakter_sceny"], p["start"], p["koniec"],
                 p["rola"], p["format"]] for p in programy],
               [16, 28, 12, 12, 12, 52, 8, 8, 16, 10])
        prog = programy
        for r, p in enumerate(prog, start=2):
            for pole, kol in (("rola", 9), ("format", 10)):
                barwa = KOLORY.get(pole, {}).get(p.get(pole, ""))
                if barwa and barwa != "FFFFFF":
                    wsp.cell(row=r, column=kol).fill = PatternFill("solid", fgColor=barwa)

    # Historia grania z RA: kto, kiedy, w jakim mieście i kraju. To jest
    # odpowiedź na „kiedy grają i w jakich warunkach" — dla całego świata,
    # nie dla trzech polskich festiwali.
    if wystepy:
        wsw = wb.create_sheet("Występy")
        arkusz(wsw, ["ksywa", "kiedy", "data", "wydarzenie", "miejsce",
                     "miasto", "kraj", "link"],
               [[w["ksywa"], w["kiedy"], w["data"], w["tytul"], w["miejsce"],
                 w["miasto"], w["kraj"], w["link"]] for w in wystepy],
               [26, 14, 12, 52, 34, 20, 22, 46])

    # KANON RA 2000-25 — jedyny w tej bazie SĄD WARTOŚCIUJĄCY. Wszystko inne
    # mówi, co się wydarzyło; ta lista mówi, co redakcja RA uznała za wzorcowe
    # w ćwierćwieczu. Osobna zakładka, bo to inny rodzaj wiedzy niż pomiar.
    if (OUT / "ra_kanon.json").exists():
        kan = json.loads((OUT / "ra_kanon.json").read_text())
        wsk = wb.create_sheet("Kanon RA")
        arkusz(wsk, ["kategoria", "miejsce", "kolejność", "ksywa", "tytuł",
                     "rok", "profil RA", "SoundCloud ID", "autor tekstu",
                     "uzasadnienie redakcji", "artykuł RA"],
               [[x["kategoria"], x["miejsce"], x["kolejnosc"], x["ksywa"],
                 x["tytul"], x["rok"], x["profil_ra"],
                 x.get("soundcloud_id") or x.get("apple_link", ""),
                 x["autor_tekstu"], x["uzasadnienie"], x["artykul"]]
                for x in kan],
               [26, 10, 12, 30, 52, 8, 40, 18, 24, 120, 46])
        # Pierwsza dziesiątka na czerwono — tylko ona ma ustalone miejsca.
        for r, x in enumerate(kan, start=2):
            if x["miejsce"]:
                wsk.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FF9999")

    # DE SCHOOL — archiwum klubu, który sam się zarchiwizował. Jedyne miejsce
    # w bazie z rozbiciem na SALE tego samego wieczoru.
    if (OUT / "de_school.json").exists():
        ds = json.loads((OUT / "de_school.json").read_text())
        wsd = wb.create_sheet("De School")
        arkusz(wsd, ["data", "dzień", "cykl", "ksywa", "scena", "typ",
                     "format", "źródło", "pewność", "strona archiwum", "link"],
               [[x["data"], x["dzien"], x["wydarzenie"], x["ksywa"], x["scena"],
                 x["typ"], x["format"], x["zrodlo"], x["pewnosc"],
                 x["strona"], x["link"]] for x in ds],
               [12, 14, 30, 30, 20, 10, 10, 26, 12, 70, 70])
        for r, x in enumerate(ds, start=2):
            for pole, kol in (("typ", 6), ("format", 7), ("pewnosc", 9)):
                barwa = KOLORY.get(pole, {}).get(x.get(pole, ""))
                if barwa and barwa != "FFFFFF":
                    wsd.cell(row=r, column=kol).fill = PatternFill("solid", fgColor=barwa)

    # TRACKLISTY — dotąd żyły wyłącznie w JSON-ie, więc w arkuszu nie było
    # widać ani jednej. Każda pozycja niesie WŁASNE `źródło`, bo komentarz
    # z SoundCloud i tracklista redakcyjna to dwa różne stopnie wiarygodności
    # i nie wolno ich mieszać w jednej kolumnie.
    if (OUT / "tracklisty_wszystkie.json").exists():
        tlw = json.loads((OUT / "tracklisty_wszystkie.json").read_text())
        wiersze_tl = []
        for w in tlw:
            for i, poz in enumerate(w["tracklista"], 1):
                wiersze_tl.append([
                    w.get("ksywa", ""), w.get("tytul", "")[:120],
                    w.get("wydarzenie", ""), w.get("data", ""),
                    w.get("polaczenie", ""), i,
                    poz.get("czas", ""),
                    "ID" if poz.get("tytul") == "ID" else "nazwany",
                    poz.get("wykonawca", ""), poz.get("tytul", ""),
                    poz.get("wydawca", ""), poz.get("zrodlo", ""),
                    w.get("link_setu", ""),
                ])
        wst = wb.create_sheet("Tracklisty")
        arkusz(wst, ["ksywa", "set", "wydarzenie", "data", "pewność połączenia",
                     "poz.", "czas", "rozpoznany", "wykonawca utworu",
                     "tytuł utworu", "wydawca", "źródło pozycji", "link setu"],
               wiersze_tl,
               [24, 46, 20, 10, 18, 6, 10, 12, 30, 42, 26, 24, 56])

    # ── SZKIELET POD ANALIZY (decyzja Janka 2026-08-14) ────────────────────
    # BPM, tonacje, czasy i szwy liczy OSOBNY tor. Tu stoją tabele z pustymi
    # kolumnami, w które te wyniki wejdą — plus wszystko, co dało się wypełnić
    # bez dotykania dźwięku: identyfikatory, zliczenia, powiązania.
    if (OUT / "encje_artysta.json").exists():
        ea = json.loads((OUT / "encje_artysta.json").read_text())
        wse = wb.create_sheet("Encje artysty")
        arkusz(wse, ["artysta_id", "nazwa kanoniczna", "ra_id", "SoundCloud",
                     "Bandcamp", "kraj", "kraj zamieszkania", "wytwórnie",
                     "obserwujących RA"],
               [[x["artysta_id"], x["nazwa_kanoniczna"], x["ra_id"],
                 x["soundcloud"], x["bandcamp"], x["kraj"],
                 x["kraj_zamieszkania"], x["wytwornie"], x["obserwujacych_ra"]]
                for x in ea],
               [12, 32, 10, 40, 44, 22, 22, 50, 16])

    if (OUT / "encje_utwor.json").exists():
        eu = json.loads((OUT / "encje_utwor.json").read_text())
        ANAL_U = ["bpm", "bpm_pewnosc", "tonacja", "tonacja_klasyczna",
                  "tonacja_pewnosc", "energia", "gestosc_groove",
                  "obecnosc_basu", "dlugosc_s", "analiza_wersja", "analiza_data"]
        wsu = wb.create_sheet("Utwory kanoniczne")
        arkusz(wsu, ["utwor_id", "wykonawca", "tytuł", "wydawca", "wystąpień",
                     "granych przez", "lata", "źródła"] + ANAL_U,
               [[x["utwor_id"], x["wykonawca"], x["tytul"], x["wydawca"],
                 x["wystapien"], x["granych_przez"], x["lata"], x["zrodla"]]
                + [x.get(k, "") for k in ANAL_U] for x in eu],
               [12, 34, 46, 28, 11, 14, 22, 34] + [10] * len(ANAL_U))
        # Puste kolumny analiz na bladym błękicie — żeby było widać na pierwszy
        # rzut oka, co jest zebrane, a co dopiero czeka na policzenie.
        czeka = PatternFill("solid", fgColor="EDF3F8")
        for kol in range(9, 9 + len(ANAL_U)):
            wsu.cell(row=1, column=kol).fill = PatternFill("solid", fgColor="4472C4")

    if (OUT / "fakty_szew.json").exists():
        fs = json.loads((OUT / "fakty_szew.json").read_text())
        ANAL_S = ["bpm_z", "bpm_do", "delta_bpm", "delta_bpm_proc", "tonacja_z",
                  "tonacja_do", "zgodnosc_harmoniczna", "dlugosc_przejscia_s",
                  "typ_przejscia", "bas_wstrzymany", "energia_z", "energia_do",
                  "delta_energii", "analiza_wersja", "analiza_data",
                  # Status w lejku i weryfikacja — dołożone po policzeniu
                  # mianownika: bez nich „492" nie dawało się umieścić
                  # w skali i nie było wiadomo, ile relacji w ogóle zgubiliśmy.
                  "zweryfikowany_przez_czlowieka", "uzywalny_do_uczenia"]
        wss = wb.create_sheet("Szwy")
        arkusz(wss, ["szew_id", "artysta_id", "ksywa", "wydarzenie", "data",
                     "poz. z", "poz. do", "utwór wychodzący", "utwór wchodzący",
                     "utwor_z_id", "utwor_do_id", "czas wejścia", "czas_ms",
                     "źródło czasu", "źródło pozycji", "status w lejku",
                     "link setu"] + ANAL_S,
               [[x["szew_id"], x["artysta_id"], x["ksywa"], x["wydarzenie"],
                 x["data"], x["pozycja_z"], x["pozycja_do"], x["utwor_z"],
                 x["utwor_do"], x["utwor_z_id"], x["utwor_do_id"], x["czas"],
                 x["czas_ms"], x["zrodlo_czasu"], x["zrodlo_pozycji"],
                 x.get("status_lejka", ""),
                 x["set_link"]] + [x.get(k, "") for k in ANAL_S] for x in fs],
               [12, 12, 24, 20, 10, 8, 8, 46, 46, 12, 12, 12, 12, 14, 22, 18, 54]
               + [10] * len(ANAL_S))
        for kol in range(18, 18 + len(ANAL_S)):
            wss.cell(row=1, column=kol).fill = PatternFill("solid", fgColor="4472C4")
        # Szwy z gotowym czasem wejścia — te da się zmierzyć od ręki.
        for r, x in enumerate(fs, start=2):
            if x["zrodlo_czasu"] == "zmierzony":
                wss.cell(row=r, column=14).fill = PatternFill("solid", fgColor="C6E0B4")
                wss.cell(row=r, column=16).fill = PatternFill("solid", fgColor="C6E0B4")

    ws4 = wb.create_sheet("Metoda")
    _art = len(pozycje)
    _mik = len(miksy)
    _pod = sum(1 for m in miksy if m.get("typ") in {"podcast", "radio", "studio"})
    _tl = 0
    if (OUT / "tracklisty_wszystkie.json").exists():
        _tl = len(json.loads((OUT / "tracklisty_wszystkie.json").read_text()))
    tekst = [
        ["STAN NA DZIEŃ BUDOWY ARKUSZA"],
        [f"Artystów: {_art}. Miksów: {_mik}, z tego podcastowych/radiowych: {_pod}."],
        [f"Tracklist zebranych z czterech źródeł: {_tl}."],
        ["Liczby liczone z danych przy każdym budowaniu — nie wpisywane ręcznie,"],
        ["bo poprzedni zestaw zdążył się zdezaktualizować w ciągu jednej sesji."],
        [""],
        ["Zakres"],
        ["Line-up 2026: Audioriver (10-12.07, Łódź), Garbicz (30.07-03.08, Torzym)"],
        ["i Wisłoujście (21-23.08, Twierdza Wisłoujście, Gdańsk — IX edycja)."],
        ["Do tego artyści,"],
        ["którzy wyszli z ARCHIWÓW obu festiwali — grali kiedyś, niekoniecznie w 2026."],
        [""],
        ["Dwa festiwale, dwie różne drogi — bo mają różną strukturę śladu"],
        ["GARBICZ ma 34 kuratorowane playlisty rocznikowe (Jeden Tag Ein Set, kutno,"],
        ["trndmsk). Wystarczyło je pobrać: 1833 pozycje jednym przebiegiem, lata 2013-2026."],
        ["AUDIORIVER nie ma ANI JEDNEJ takiej kolekcji — sety leżą pojedynczo, wrzucane"],
        ["przez samych artystów. Jedyne wejście hurtowe to wyszukiwarka SoundCloud"],
        ["(api-v2), stronicowana. Stąd asymetria: 744 sety z Garbicza, 184 z Audioriver."],
        ["To NIE znaczy, że Audioriver jest mniejszy. Znaczy, że jego publiczność nie"],
        ["prowadzi archiwum — a to samo w sobie jest obserwacją o obu festiwalach."],
        [""],
        ["WISŁOUJŚCIE — jedyny festiwal z PEŁNYM programem"],
        ["Strona ładuje wszystko javascriptem, ale js/lineup.js i js/timetable.js sięgają"],
        ["po dwa zwykłe pliki: data/lineup.json i data/schedule.json. Drugi to pełny"],
        ["timetable z godzinami start-koniec, per scena, per dzień. Stąd osobna zakładka."],
        ["Dwie rzeczy, które odróżniają go od tamtych: line-up jest WYŁĄCZNIE POLSKI"],
        ["(deklaracja organizatora), a SCENA NIESIE GATUNEK — Twierdza to techno,"],
        ["Szaniec tech-house i disco, Raj downtempo, Bastion industrial i rave."],
        ["Program to PLAN, nie nagrania: festiwal zagra dopiero 21 sierpnia. Dlatego"],
        ["siedzi osobno od arkusza Miksy, gdzie leży to, czego da się posłuchać."],
        ["`rola` jest tu PEWNA, nie wyłuskana z opisu — wynika z pozycji w timetable."],
        [""],
        ["Co znaczy puste pole"],
        ["Puste = nie znalazłam. Nigdy 'nie istnieje' i nigdy nie zgadywane."],
        [""],
        ["Wiersze podświetlone na żółto"],
        ["Apple Music zwrócił artystę o tej samej nazwie, ale z gatunku spoza muzyki"],
        ["klubowej (K-Pop, Metal, Anime). To prawie na pewno inny wykonawca — link"],
        ["zostaje do ręcznego sprawdzenia, a nie jako fakt."],
        [""],
        ["Imię i nazwisko"],
        ["Kolumny nie ma — decyzja Janka. Dane osobowe, których artyści sami nie podają."],
        [""],
        ["Utwory"],
        ["Źródło: iTunes Search API, katalog Apple Music, kolejność zwracana przez API."],
        ["NIE jest to potwierdzona lista najpopularniejszych — API nie podaje odtworzeń."],
        [""],
        ["Katalog utworów: DWA źródła, bo żadne nie wystarcza"],
        ["Apple Music (iTunes Search API) i Bandcamp (wyszukiwarka bcsearch_public_api)."],
        ["Nakładka jest niepełna z obu stron: część artystów zna tylko Apple, część tylko"],
        ["Bandcamp. Przy line-upie Wisłoujścia — wyłącznie polskim i undergroundowym —"],
        ["Apple trafiło 14%, bo ci ludzie wydają na Bandcampie i winylu. Puste pole"],
        ["w jednym serwisie nie znaczy brak muzyki, tylko brak w TYM serwisie."],
        ["Bandcamp daje przy okazji LOKALIZACJĘ (wpisaną ręką artysty) i flagę"],
        ["'to wytwórnia, nie artysta' — jedyne miejsce, gdzie serwis mówi to sam."],
        [""],
        ["Tracklisty — cztery źródła, trzy stopnie pewności połączenia"],
        ["SoundCloud (komentarze przypięte do sekundy nagrania), MixesDB, NTS Radio,"],
        ["hearthis.at. Liczba pozycji NIE jest miarą wartości — miarą jest to, czy"],
        ["wiemy, DO KTÓREGO SETU należą. Stopnie: 'link' (ten sam adres SoundCloud"],
        ["po obu stronach — fakt), 'tytul+rok' (mocna poszlaka, tylko gdy trafia"],
        ["w jeden wiersz), 'nowy' (tracklista bez naszego setu). Nie łączymy po samej"],
        ["ksywie: 'Ben Klock @ Berghain 2019' i '@ Garbicz 2019' to dwa różne sety."],
        [""],
        ["Rezydencja i afiliacja — kolumny USUNIĘTE 2026-08-14"],
        ["Decyzja Janka. Jedyne pole bez taniego źródła: 24 wypełnione na 1007, czyli"],
        ["kolumna głównie ogłaszała, czego nie wiemy. Zebrane wartości nie przepadły —"],
        ["siedzą w socials.json i wrócą, jeśli znajdzie się sposób na nie hurtem."],
        [""],
        ["Miksy: skąd pochodzą"],
        ["1001Tracklists — baza indeksująca WYŁĄCZNIE realne sety DJ-skie, z miejscem i datą."],
        ["To jest źródło pewne: fanowskie składanki tam nie istnieją, a artyści o tej samej"],
        ["nazwie mają osobne strony (novah-(be) vs novah-(sk)), więc baza sama je rozróżnia."],
        ["YouTube — daje link do słuchania, ale wymaga oceny: odrzucałam pliki typu"],
        ["'Best Of HARD TECHNO 2026 | NOVAH, CHARLOTTE DE WITTE', bo to składanki fanowskie"],
        ["z cudzych utworów, nie sety tych artystów."],
        [""],
        ["Archiwum Garbicza 2013-2026"],
        ["34 playlisty rocznikowe ze SoundCloud (Jeden Tag Ein Set, kutno, trndmsk i inne)"],
        ["plus dobicie wyszukiwarką api-v2, która wyciągnęła sety spoza kolekcji."],
        ["Zaciągnięte hurtem, nie po jednym secie — strona SoundCloud wstrzykuje pełny opis"],
        ["playlisty do HTML-a, więc jedno pobranie daje kilkadziesiąt setów naraz."],
        ["Kolekcja 'Garbicz Lineup 2024' jest ZAPOWIEDZIĄ line-upu, nie zapisem festiwalu:"],
        ["106 ze 198 pozycji to sety tych artystów zagrane gdzie indziej. Odsiane — wiersz"],
        ["trafia do Garbicza tylko gdy pada nazwa festiwalu albo parkiet występujący"],
        ["wyłącznie tam (Seebühne, Buk Corner, Juicy Bar, Crazy Paradise, Lichtung)."],
        [""],
        ["Kolumna 'kto grał obok'"],
        ["Odpowiada na pytanie z samego środka DanceLab: kto po kim. Wypełniana WYŁĄCZNIE"],
        ["ręcznie. Automat na wzorcu 'after X' dawał 60% fałszywek — '2 days before"],
        ["Garbicz.... Panic!' czytał jako sąsiada. Zostało 7 wierszy,"],
        ["w których artysta napisał to wprost. Siedem pewnych bije czterdzieści zmyślonych."],
        [""],
        ["Kolumna 'opis od artysty'"],
        ["Tekst, który artysta sam napisał pod setem. Stąd pochodzi większość godzin,"],
        ["dni tygodnia i ról — a także rzeczy, których nie da się wyliczyć: kto grał przed"],
        ["kim, jaka była pogoda, czy to była ostatnia godzina festiwalu."],
        [""],
        ["Format 'live' — czego NIE wpisujemy"],
        ["Słowo 'live' w tytule znaczy dwie różne rzeczy: 'gram na maszynach' i 'to jest"],
        ["nagranie z imprezy'. Rozstrzyga pozycja: doklejone do nazwy artysty ('SKINNERBOX"],
        ["Live @ Garbicz') to format. Na początku tytułu ('Live@ Garbicz 2014') opisuje"],
        ["nagranie i pole zostaje PUSTE — 50 wierszy tak ma."],
    ]
    for w in tekst:
        ws4.append(w)
    ws4.column_dimensions["A"].width = 100
    for r, w in enumerate(tekst, start=1):
        if w and w[0] and not w[0].startswith(("Puste", "456", "Audioriver", "Apple",
                                               "klubowej", "zostaje", "Kolumny",
                                               "Źródło", "NIE", "Z oficjalnego")):
            ws4.cell(row=r, column=1).font = Font(bold=True)

    ws5 = wb.create_sheet("Legenda")
    ws5.append(["wymiar", "wartość", "kolor", "co znaczy"])
    for c in ws5[1]:
        c.font, c.fill = naglowek, tlo
    OPISY = {
        "festiwal": "impreza plenerowa albo wielosceniczna",
        "klub": "lokal z parkietem",
        "warehouse": "hala, magazyn, przestrzeń tymczasowa",
        "rave": "impreza w miejscu nieoczywistym",
        "plener": "na zewnątrz, poza festiwalem",
        "radio": "stacja albo audycja radiowa",
        "studio": "nagranie bez publiczności (HÖR, The Lab, Boiler Room studyjny)",
        "podcast": "cykl wydawnictwa albo medium",
        "stream": "transmisja bez publiczności",
        "dj-set": "domyślny — gra z płyt lub plików",
        "live": "gra z instrumentów albo maszyn",
        "b2b": "dwoje lub więcej na zmianę",
        "winyl": "set deklarowany jako tylko z płyt",
        "otwarcie": "pierwszy set sceny albo festiwalu",
        "peak": "szczyt wieczoru",
        "zamkniecie": "ostatni set sceny albo festiwalu",
        "afterhour": "po zamknięciu głównego programu",
        "poranek": "set poranny",
        "wschod-slonca": "deklarowany jako wschód słońca",
        "popoludnie": "set popołudniowy (afternoon, Nachmittag, day time)",
        "zachod-slonca": "sunset albo sundowner",
        "noc": "set nocny bez bliższego określenia",
        "all-night": "jeden artysta przez całą noc",
        "1001tracklists": "baza z datą i miejscem — źródło najpewniejsze",
        "soundcloud": "wrzut artysty — pewny, ale bez ustandaryzowanych metadanych",
        "youtube": "wymaga oceny — trafiają się składanki fanowskie",
        "potwierdzone": "wiem, że to ten artysta",
        "niepewne": "mam kandydata, nie potwierdziłem",
    }
    wiersz = 2
    for wymiar, mapa in KOLORY.items():
        for wart, barwa in mapa.items():
            ws5.cell(row=wiersz, column=1, value=wymiar)
            ws5.cell(row=wiersz, column=2, value=wart)
            ws5.cell(row=wiersz, column=3).fill = PatternFill("solid", fgColor=barwa)
            ws5.cell(row=wiersz, column=4, value=OPISY.get(wart, ""))
            wiersz += 1
        wiersz += 1
    for lit, sz in zip("ABCD", (16, 20, 10, 70)):
        ws5.column_dimensions[lit].width = sz
    ws5.freeze_panes = "A2"

    plik = OUT / "mapa_djow_audioriver_garbicz.xlsx"
    wb.save(plik)

    # Mianownikiem jest cała tabela Artyści, nie sam line-up 2026 — od czasu
    # zaciągnięcia roczników Garbicza to dwie różne liczby i mylenie ich
    # zawyżałoby pokrycie o ponad połowę.
    n = len(pozycje)
    apl = sum(1 for r in ap.values() if r.get("apple_music"))
    pod = sum(1 for r in ap.values() if r.get("do_sprawdzenia"))
    print(f"zapisane: {plik}\n")
    print(f"POKRYCIE na {n} artystach "
          f"({len(fest)} z line-upu 2026 + {len(z_rocznikow)} z archiwów obu festiwali):")
    print(f"  Apple Music (dopasowanie po nazwie)  {apl:3d}/{n}  ({apl / n * 100:.0f}%)")
    print(f"    z tego prawdopodobnie inny wykonawca  {pod:3d}")
    print(f"    czyli wiarygodnych                    {apl - pod:3d}/{n}  "
          f"({(apl - pod) / n * 100:.0f}%)")
    print(f"  utwory                               {len(apple['utwory']):3d} pozycji")
    sc_n = sum(1 for v in soc.values() if v.get("soundcloud")) + sum(
        1 for v in RECZNIE.values() if v[1])
    kand_n = sum(1 for v in soc.values() if v.get("kandydaci"))
    print(f"  SoundCloud                           {sc_n:3d}/{n}  ({sc_n / n * 100:.0f}%)")
    print(f"  wierszy z kandydatami do wyboru      {kand_n:3d}")
    print(f"  miksów w bazie                      {len(miksy):4d}")
    print(f"    z tego setów z Garbicza           "
          f"{sum(1 for m in miksy if 'garbicz' in (m.get('wydarzenie') or '').lower()):4d}")
    print(f"    z opisem od artysty               "
          f"{sum(1 for m in miksy if (m.get('opis') or '').strip()):4d}")
    print(f"    z wypełnionym 'kto grał obok'     "
          f"{sum(1 for m in miksy if m.get('sasiedztwo')):4d}")
    import collections as _c
    pod = _c.Counter(m["ksywa"] for m in miksy
                     if m.get("typ") in {"podcast", "radio", "studio"} and m.get("ksywa"))
    print(f"    setów podcastowych/radiowych      "
          f"{sum(1 for m in miksy if m.get('typ') in {'podcast','radio','studio'}):4d}")
    print(f"      artystów z >=1                  {len(pod):4d}")
    print(f"      artystów z >=3 (cel Janka)      "
          f"{sum(1 for v in pod.values() if v >= 3):4d}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
