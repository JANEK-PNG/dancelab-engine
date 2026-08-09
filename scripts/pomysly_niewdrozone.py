"""Wszystkie pomysły Janka, które NIE zostały wdrożone → arkusz na biurko.

Użycie:
    .venv/bin/python scripts/pomysly_niewdrozone.py

Źródła: pamięć projektu (49 plików), PROJECT_LEDGER.md (sekcje ZAPARKOWANE
i PYTANIA DO JANKA + dziennik), docs/ w repo silnika, bieżąca sesja.

Ocena wg skali Janka: KILLER / HIGH / MID / LOW / CRITICAL (nie wdrażać).
CRITICAL nie znaczy „ważne" — znaczy „zmierzone i odrzucone, nie wracać".
"""
import pathlib

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # jedyna zależność, celowo spoza wymagań silnika
    print("brakuje openpyxl — zainstaluj:\n"
          "  .venv/bin/pip install openpyxl")
    raise SystemExit(1)

WYJSCIE = pathlib.Path.home() / "Desktop" / "DanceLab — pomysły niewdrożone.xlsx"

KOLORY = {
    "KILLER":   ("FF6D00", "FFFFFF"),
    "HIGH":     ("FFD54F", "1A1A1A"),
    "MID":      ("C5E1A5", "1A1A1A"),
    "LOW":      ("E0E0E0", "1A1A1A"),
    "CRITICAL": ("C62828", "FFFFFF"),
}

# (pomysł, co to jest, skąd i kiedy, stan dzisiaj, ocena, dlaczego taka ocena,
#  co trzeba zrobić / co odblokowuje)
POMYSLY = [
    # ---------------------------------------------------------------- KILLER
    ("Silnik uczy się CIEBIE z Twojej historii",
     "Zamiast jednego silnika dla wszystkich — DanceLab czyta Twoją historię "
     "grania z Rekordboxa i uczy się Twoich wyborów. Lokalnie, w minuty, bez "
     "wysyłania czegokolwiek.",
     "Twoja prośba 01.08: „przeczytaj o ML i powiedz, jak zrobić z DanceLaba "
     "maszynę uczącą się, wszystko lokalnie”",
     "ZERO kodu. Rozpisana architektura (3 piętra) i 5 pytań do Ciebie "
     "(#11–#15 w rejestrze) czeka na decyzję.",
     "KILLER",
     "To jedyna rzecz, która usuwa zmierzoną dziurę: w dzisiejszej ocenie "
     "przejść NIE MA NIC z tego, co wiemy o Tobie. Reszta pomysłów to "
     "ulepszenia, ten zmienia produkt z „kalkulatora” w „Twojego”.",
     "Twoja decyzja na pytania #11–#15. Warunek techniczny: podział danych "
     "po sesjach (GroupKFold), inaczej wynik będzie zawyżony."),

    ("Werdykty wracają do oceny (pętla uczenia)",
     "Każda Twoja poprawka setu (cięcie, podmiana, przesunięcie) już się "
     "zapisuje — ale silnik ich NIE CZYTA. Ten pomysł zamyka pętlę: im "
     "więcej używasz, tym lepiej dobiera.",
     "Wynika wprost z ML (pytanie #14, 01.08)",
     "Werdykty zapisywane od tygodni. Konsumenta brak — leżą i się kurzą.",
     "KILLER",
     "Najtańszy krok o największym skutku: dane JUŻ się zbierają, brakuje "
     "tylko czytania. Bez tego używanie programu niczego go nie uczy.",
     "Decyzja jak wyżej. Uwaga: dane zbierają się tylko do przodu, więc "
     "każdy tydzień zwłoki to tydzień nienauczony."),

    ("Lista do kopania (digging) w aplikacji",
     "Utwory z setów prawdziwych DJ-ów, najbliższe Twojej bibliotece, "
     "których NIE MASZ — każdy z wyjaśnieniem „bo brzmi jak twój X”.",
     "Twój pomysł, zbudowany 23.07",
     "DZIAŁA, ale jako skrypt obok produktu (`scripts/digging_list.py`). "
     "W aplikacji tego nie ma.",
     "KILLER",
     "Jedyna funkcja, która daje wartość ZANIM zbudujesz set, i sama się "
     "obroniła: znalazła kolejne utwory artystów, których masz, wyłącznie "
     "po brzmieniu. Wersja z kotwicą dawała czysty leftfield.",
     "Wpiąć jako zakładkę albo panel w TUI. Kod i dane gotowe."),

    # ------------------------------------------------------------------ HIGH
    ("Perfect Duo — kotwica z dwóch DJ-ów",
     "Wybierasz dwóch DJ-ów i DanceLab układa set tak, jakby grali b2b. "
     "Każdy utwór z inicjałami tego, do kogo brzmienia został dobrany.",
     "Twój pomysł 09.08 (dzisiaj)",
     "Zapisany, nie wdrożony — Twoja decyzja.",
     "HIGH",
     "Tanie w budowie, bo rodziny brzmieniowe i podobieństwo już liczymy; "
     "a produktowo to jedyny pomysł, w którym kotwica przestaje być filtrem, "
     "a staje się dwoma głosami.",
     "Rozstrzygnąć: b2b zmienia się co utwór, co fragment, czy blokami."),

    ("Wstępne sito brzmieniowe przy dużych pulach",
     "Przy tysiącach kandydatów najpierw odsiać po brzmieniu (trzy zgrubne "
     "cechy), potem liczyć dokładnie.",
     "Wniosek z badania tripletów 08–09.08",
     "Niewdrożone. Zmierzone, że bez tego most-triplet gubi przewagę przy "
     "wielkich pulach.",
     "HIGH",
     "Odblokowuje jedyną rzecz, którą zmierzyliśmy jako REALNY zysk "
     "(+8–15 punktów procentowych przy ustalonym filarze).",
     "Wpiąć CLAP jako filtr wstępny w budowie i przy podpowiedziach."),

    ("Rozkład szwu — jak DJ to zmiksował",
     "Mając miks i oba utwory, odzyskać ruchy rąk: fader, EQ, długość "
     "przejścia. Odwrotność naszego automiksu.",
     "Twój pomysł, korpus v2 (17.07)",
     "Częściowo: profil szwu policzony na Twoich nagraniach (171 uderzeń "
     "blendu, bas wstrzymany w 86% wejść). Pełnego rozkładu brak.",
     "HIGH",
     "To jest fosa produktu — nikt inny nie mierzy SZWU, tylko utwory. "
     "Reszta rynku dobiera tracki, my potrafilibyśmy opisać ruch.",
     "Uwaga zmierzona: naiwne odejmowanie (miks − A − B) NIE działa; "
     "trzeba rozwiązywać układ z EQ i gainami."),

    ("Program działa u innych, nie tylko u Ciebie",
     "Podpis Apple (Developer ID + notarization) i test na czystym Macu, "
     "żeby dało się rozdać kolegom bez ostrzeżeń systemu.",
     "Audyt 23.07, pytanie #7 do Ciebie",
     "Niezrobione. Dziś aplikacja realnie działa tylko na Twoim komputerze.",
     "HIGH",
     "DanceLab Wrapped budowaliśmy PO TO, żeby rozdawać go kolegom-DJ-om. "
     "Bez podpisu ten cel jest zablokowany na starcie.",
     "Koszt 99 USD rocznie + jednorazowa konfiguracja. Twoja decyzja."),

    ("Tablica inspiracji „graj jak mistrzowie”",
     "Karty DJ-ów jak kolekcjonerskie: wybierasz swoją talię, a program "
     "podpowiada zestawienia i przejścia w ich duchu.",
     "Twój pomysł 16.07",
     "Częściowo: kotwica „graj jak…” i rodziny brzmieniowe DZIAŁAJĄ. "
     "Kart i talii nie ma.",
     "HIGH",
     "Połowa jest już zbudowana i sprawdzona; brakuje warstwy, która "
     "z narzędzia robi coś, co chce się pokazać innym.",
     "Wymaga wpięcia centroidów mistrzów jako biasu selekcji."),

    # ------------------------------------------------------------------- MID
    ("Enter zatwierdza we wszystkich panelach",
     "Dziś Enter działa w gatunkach i kotwicach. W podmianie (Z), dopisaniu "
     "(A) i liście planów (O) nadal trzeba nacisnąć skrót drugi raz.",
     "Twoja uwaga 09.08 o Ctrl+G — ten sam problem",
     "Niewdrożone, czeka na Twoje „tak”.",
     "MID",
     "Spójność: albo Enter znaczy „zatwierdzam” wszędzie, albo nigdzie. "
     "Dziś znaczy w połowie miejsc.",
     "Pół godziny pracy. Powiedz i robię."),

    ("Graf artystów (Last.fm)",
     "Mapa pokrewieństwa artystów z zewnętrznego źródła — uzupełnia to, "
     "czego nie słychać w samym dźwięku.",
     "Zaparkowane, pytanie #4 do Ciebie",
     "Crawler GOTOWY. Brakuje darmowego klucza API — dwie minuty Twojego "
     "czasu.",
     "MID",
     "Wszystko zrobione poza jednym kliknięciem po Twojej stronie; szkoda, "
     "żeby leżało.",
     "Załóż klucz na last.fm i wklej."),

    ("Prior czasu na deckach (dwell time)",
     "Ile realnie utwór gra u Ciebie, zanim wejdzie następny — jako kolejna "
     "przesłanka przy doborze.",
     "Rozmowa głosowa 08.08",
     "Niewdrożone.",
     "MID",
     "Dane są (Twoje nagrane sety), a to prosty i konkretny sygnał o Twoim "
     "sposobie grania.",
     "Policzyć z nagrań i wpiąć jako prior."),

    ("Korpus v2 — więcej setów (1001Tracklists)",
     "Poszerzenie zbioru realnych setów, na których mierzymy przejścia.",
     "Zaparkowane 17.07",
     "Backlog.",
     "MID",
     "Więcej danych = pewniejsze liczby, ale dzisiejszy korpus jeszcze nie "
     "jest wyciśnięty do końca.",
     "Dokończyć obecny korpus, potem to."),

    ("Pokój AI — Klaris i Kord rozmawiają, Ty patrzysz",
     "Broker do rozmowy dwóch asystentów; Ty za lustrem, treść spotkania "
     "nigdy na wierzchu.",
     "Twój projekt, stan spisany 15.07",
     "Kod stoi w Desktop/AI/ai-room. Inauguracja nieodpalona.",
     "MID",
     "Ciekawe narzędzie pracy, ale nie posuwa DanceLaba do przodu.",
     "Odpalić inaugurację albo świadomie zamknąć."),

    ("Reszta poprawek nazewnictwa w interfejsie",
     "Dziewięć drobiazgów z audytu: skok/uderzenie, PODMIANA/Zamień, "
     "wielkie litery, „backupy”, „pula” w dwóch znaczeniach, "
     "„deterministyczny”, kolumna „pew.”, Z jako cofnij i zamień naraz.",
     "Audyt nazewnictwa 09.08 — zrobiliśmy pierwszą piątkę",
     "Niewdrożone.",
     "MID",
     "Każdy z osobna jest drobiazgiem, razem to różnica między „narzędziem "
     "dla mnie” a „narzędziem dla kogokolwiek”.",
     "Jedna sesja."),

    ("Apple Music — dojście do człowieka",
     "Publikacja (ISMIR Late-Breaking Demo) i dotarcie do żywej osoby "
     "zamiast systemu rekrutacyjnego.",
     "Twoja ścieżka kandydata",
     "CV, esej i notatki gotowe. Publikacji i kontaktu brak.",
     "MID",
     "Poza produktem, ale to Twój cel zawodowy i ma termin.",
     "Wybrać konferencję i napisać dwie strony."),

    # ------------------------------------------------------------------- LOW
    ("Brief setu dyktowany po ludzku (model językowy)",
     "Mówisz „set na łódkę o zachodzie, wchodzę po kimś szybszym”, a system "
     "tłumaczy to na parametry.",
     "Twoja korekta 27.07",
     "Niewdrożone.",
     "LOW",
     "Twoje własne słowa: priorytet niski względem maszyny od szwów. "
     "Do tego brief da się dziś wyklikać w kilkanaście sekund.",
     "Najpierw rozstrzygnąć, czy to wejście, wyjście, czy oba."),

    ("Portrety DJ-ów z danych („PRZEJŚCIA”)",
     "Obraz DJ-a zbudowany wyłącznie z jego zmierzonych przejść — pierścień "
     "temp, łuki długości blendów, ruch harmoniczny jako barwa.",
     "Twoje „Przejścia brzmi genialnie”, 16.07",
     "Koncept i gramatyka wizualna spisane. Zero kodu.",
     "LOW",
     "Piękny projekt-satelita, ale nie zmienia ani jednego setu.",
     "Po pełnej analizie korpusu."),

    ("Karty stylu 551 DJ-ów (DJ Style Dex)",
     "Wizualne karty wszystkich zebranych DJ-ów.",
     "Przerwane w połowie budowy",
     "Niedokończone.",
     "LOW",
     "Ładne, ale nakłada się z tablicą inspiracji, która jest wyżej.",
     "Jedna sesja, jeśli w ogóle."),

    ("Muzyka generatywna",
     "Dwanaście utworów z czystego DSP, wzór Kordiego, faza i rama.",
     "Twoja decyzja 01.08: wydzielić z rdzenia",
     "Osobny wątek, świadomie odłożony.",
     "LOW",
     "Ty sam to zaparkowałeś, żeby nie rozmywać DanceLaba przed testami.",
     "Osobna sesja, gdy będzie ochota."),

    ("CURVE — tryb offline i skala energii 1–10",
     "Biblioteka i sety bez serwera; skala energii zgodna z Mixed In Key.",
     "Odłożone 06.07",
     "Projekt uśpiony od lipca.",
     "LOW",
     "CURVE rozwiązuje ten sam problem od strony interfejsu, a DanceLab "
     "rozwiązuje go od strony pomiaru — i to DanceLab wygrał Twój czas.",
     "Nie ruszać, dopóki DanceLab nie stanie."),

    ("Koszyk Bandcamp z setu Four Teta",
     "Lista utworów z konkretnego setu wrzucona do koszyka.",
     "Twoja prośba 08.08",
     "Zablokowane: wtyczka przeglądarki niepodłączona; 31 z 54 utworów ma "
     "linki.",
     "LOW",
     "Jednorazowa wygoda, nie funkcja produktu.",
     "Podłączyć wtyczkę albo zrobić ręcznie z gotowej listy."),

    ("Badanie z pięcioma oceniającymi",
     "Sprawdzić silnik na ludziach spoza jednej głowy.",
     "Zaparkowane",
     "Niezrobione.",
     "LOW",
     "Ma sens dopiero, gdy silnik przestanie się zmieniać co dwa dni.",
     "Po ustabilizowaniu wag."),

    ("Migotliwy test interfejsu",
     "Jeden test TUI potrafi zamrugać raz na kilkanaście przebiegów.",
     "Obserwacja 09.08",
     "Nietropiony.",
     "LOW",
     "Nie psuje programu, ale zaciemnia obraz przy każdym pełnym przebiegu.",
     "Godzina z pytest-randomly."),

    # -------------------------------------------------- CRITICAL (nie wdrażać)
    ("Cała aplikacja na sieci neuronowej",
     "Zastąpić liczenie jedną dużą siecią, która „sama się wszystkiego "
     "nauczy”.",
     "Twój pomysł 01.08, sprawdzony i odradzony",
     "Odrzucone po pomiarze.",
     "CRITICAL",
     "Sieć ZAWSZE coś zwróci — nie umie powiedzieć „nie wiem”, a to jest "
     "jedyna kompetencja, której konkurencja nie ma. Do tego arytmetyka: "
     "~2345 par następstwa to sufit rzędu 50–200 parametrów. Sieć trenowana "
     "przez nas jest wykluczona danymi, nie ambicją.",
     "Zamiast tego: cienka warstwa ucząca się nad nietykalnym pomiarem "
     "(pozycja 1 tej listy)."),

    ("Grupowanie DJ-ów po gatunku albo po klubie",
     "Rodziny DJ-ów budowane z etykiet gatunkowych lub z miejsc, w których "
     "grali.",
     "Rozważane przy „Graj jak…”",
     "Odrzucone po sprawdzeniu na danych.",
     "CRITICAL",
     "Pole gatunku jest PUSTE w 100% z 551 profili — dopisanie go 284 "
     "nazwiskom byłoby zgadywaniem. Miejsce (Boiler Room, Warehouse "
     "Project) to nazwa NASZEGO źródła, nie tożsamość artysty; grupowanie "
     "po tym kłamałoby.",
     "Zostaje to, co działa: rodziny liczone z brzmienia nagrań."),

    ("Ślepe wykrywanie szwów w miksie",
     "Znajdować przejścia analizując sam miks, bez utworów źródłowych.",
     "Eksperyment 05",
     "Zmierzone i odrzucone.",
     "CRITICAL",
     "Na prawdziwym secie skuteczność wyszła 0,26 — to jest rzut monetą "
     "z gorszym humorem. Droga, która działa, to dopasowanie do utworów "
     "źródłowych.",
     "Nie wracać bez zupełnie nowej metody."),

    ("Rozkład szwu przez proste odejmowanie",
     "Policzyć ruchy DJ-a jako „miks minus utwór A minus utwór B”.",
     "Pierwsza intuicja przy rozkładzie szwu",
     "Odrzucone rachunkiem.",
     "CRITICAL",
     "Miks to nie suma — to dwa tory ze zmiennym w czasie faderem, "
     "nieznanym wzmocnieniem i trzypasmowym EQ. Odejmowanie daje śmieci "
     "i wygląda przekonująco, co jest najgorszą kombinacją.",
     "Rozwiązywać układ, nie odejmować."),

    ("Zamiana CLAP na MuQ",
     "Podmienić „słuch maszyny” na nowszy model.",
     "Eksperyment 01 (22.07)",
     "Przetestowane, odrzucone.",
     "CRITICAL",
     "Zero przewagi w teście brzmieniowym (11 do 10), a MuQ ma huby — "
     "utwory bliskie wszystkiemu. Do tego licencja niekomercyjna, więc "
     "produktowo ślepa uliczka.",
     "CLAP zostaje. Następcy szukać dopiero, gdy pojawi się mocny."),

    ("Kopiowanie tempa i siatki z Rekordboxa",
     "Brać BPM i beatgrid z Pioneera zamiast liczyć samemu.",
     "Kusiło przy sporach o tempo",
     "Zakazane twardo.",
     "CRITICAL",
     "Eksport NIGDY nie rusza tempa ani siatki — to jest niepodważalna "
     "zasada projektu; jedno pomyłkowe nadpisanie psuje Twoją bibliotekę "
     "bez odwrotu. Przy sporze pokazujemy OSTRZEŻENIE, nie podmieniamy.",
     "Zostaje `tempo_disputes`: mówimy, że się nie zgadzamy."),
]

KOLEJNOSC = {"KILLER": 0, "HIGH": 1, "MID": 2, "LOW": 3, "CRITICAL": 4}
NAGLOWKI = ["#", "Pomysł", "Co to jest", "Skąd i kiedy", "Stan dzisiaj",
            "Ocena", "Dlaczego taka ocena", "Co trzeba zrobić"]
SZEROKOSC = [4, 34, 52, 30, 34, 12, 56, 38]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pomysły"

    ws["A1"] = "DanceLab — pomysły Janka, które NIE zostały wdrożone"
    ws["A1"].font = Font(size=15, bold=True)
    ws["A2"] = ("Zebrane 2026-08-09 z pamięci projektu, rejestru "
                "PROJECT_LEDGER.md i dokumentów w repozytorium silnika. "
                "CRITICAL nie znaczy \u201eważne\u201d \u2014 znaczy "
                "\u201esprawdzone i odrzucone, nie wracać\u201d.")
    ws["A2"].font = Font(size=10, italic=True, color="666666")
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    wiersz_naglowka = 4
    for i, tekst in enumerate(NAGLOWKI, start=1):
        k = ws.cell(row=wiersz_naglowka, column=i, value=tekst)
        k.font = Font(bold=True, color="FFFFFF")
        k.fill = PatternFill("solid", fgColor="37474F")
        k.alignment = Alignment(vertical="center")
    ws.row_dimensions[wiersz_naglowka].height = 22

    kreska = Side(style="thin", color="DDDDDD")
    obwodka = Border(bottom=kreska)

    dane = sorted(POMYSLY, key=lambda p: KOLEJNOSC[p[4]])
    for n, (nazwa, opis, skad, stan, ocena, czemu, co) in enumerate(dane, 1):
        r = wiersz_naglowka + n
        wartosci = [n, nazwa, opis, skad, stan, ocena, czemu, co]
        for i, wartosc in enumerate(wartosci, start=1):
            k = ws.cell(row=r, column=i, value=wartosc)
            k.alignment = Alignment(wrap_text=True, vertical="top")
            k.border = obwodka
            if i == 2:
                k.font = Font(bold=True)
        tlo, tekst = KOLORY[ocena]
        k = ws.cell(row=r, column=6)
        k.fill = PatternFill("solid", fgColor=tlo)
        k.font = Font(bold=True, color=tekst)
        k.alignment = Alignment(horizontal="center", vertical="center")
        # BEZ sztywnej wysokości — Excel i Numbers dopasują wiersz do
        # zawiniętego tekstu same; ustawiona na sztywno ucinała dłuższe
        # uzasadnienia (złapane na podglądzie arkusza)

    for i, szer in enumerate(SZEROKOSC, start=1):
        ws.column_dimensions[get_column_letter(i)].width = szer
    ws.freeze_panes = f"A{wiersz_naglowka + 1}"
    ws.auto_filter.ref = (f"A{wiersz_naglowka}:H{wiersz_naglowka + len(dane)}")

    # ------------------------------------------------------ druga karta: skala
    ws2 = wb.create_sheet("Skala ocen")
    ws2["A1"] = "Co znaczy która ocena"
    ws2["A1"].font = Font(size=14, bold=True)
    opisy = [
        ("KILLER", "Zmienia produkt, nie ulepsza go. Bez tego DanceLab "
                   "zostaje kalkulatorem."),
        ("HIGH", "Duża wartość, podstawy już zbudowane — brakuje wpięcia "
                 "albo jednej decyzji."),
        ("MID", "Warto zrobić, ale świat się nie zawali. Zwykle tanie."),
        ("LOW", "Miłe, poboczne albo świadomie przez Ciebie odłożone."),
        ("CRITICAL", "NIE WDRAŻAĆ. Sprawdzone i odrzucone pomiarem — wpisane "
                     "tu po to, żeby nie wracać do tego za pół roku."),
    ]
    for i, (ocena, opis) in enumerate(opisy, start=3):
        k = ws2.cell(row=i, column=1, value=ocena)
        tlo, tekst = KOLORY[ocena]
        k.fill = PatternFill("solid", fgColor=tlo)
        k.font = Font(bold=True, color=tekst)
        k.alignment = Alignment(horizontal="center")
        o = ws2.cell(row=i, column=2, value=opis)
        o.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[i].height = 34
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 80

    licz = {o: sum(1 for p in POMYSLY if p[4] == o) for o in KOLORY}
    ws2["A10"] = "Ile czego"
    ws2["A10"].font = Font(size=12, bold=True)
    for i, (ocena, ile) in enumerate(
            sorted(licz.items(), key=lambda x: KOLEJNOSC[x[0]]), start=11):
        ws2.cell(row=i, column=1, value=ocena)
        ws2.cell(row=i, column=2, value=f"{ile} pozycji")

    WYJSCIE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WYJSCIE)
    print(f"{WYJSCIE}  ({WYJSCIE.stat().st_size / 1024:.0f} KB)")
    print(f"pozycji: {len(POMYSLY)} · " + " · ".join(
        f"{o}: {n}" for o, n in sorted(licz.items(),
                                       key=lambda x: KOLEJNOSC[x[0]])))


if __name__ == "__main__":
    main()
